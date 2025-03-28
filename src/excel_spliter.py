"""
@python version:
    Python 3.4

@summary:
    Control composed by the ExcelSplitter() class.
    This class has:
        - constructor:
            - __init__(excel_path);
        - a public function/ method:
            - splitter()
        - five non-public / helper function / methods:
            - __create_temp_folder()
            - __excel_lines_counter()
            - __excel_titles_row()
            - __create_temp_excel(start_row, end_row, temp_number, titles_row)
            _ __divider(num_rows, divisor=2) 

@note:
    function __init__ (excel_path)
        Class constructor
        excel_path is a string of the path to the excel file
        initialises the class variables AND constant:
            - excel_path - the path to the excel file
            - folder - the path to the folder who contains the excel file
            - file - the file name of the excel file
            - MAX_NUM_ROWS - a limiter for the maximum number of rows for an excel file,
                initialised to 5000 (constant)

    function splitter()
        Verifies if the first four columns of the first row are correct
        Check if the Excel file has more than 5000 rows, divide it if yes and returns
        the a list of strings of the paths of the Excel files

    function __create_temp_folder()
        Makes a Temp folder child of the current folder

    function  __excel_lines_counter()
        Counts and returns the number of rows of all sheets in an Excel file

    function __excel_titles_row()
        Builds and returns a list containing the first row of all sheets in an Excel file

    function __create_temp_excel(start_row, end_row, temp_number, titles_row)
        Creates a new Excel file with no more than 5000 rows in the Temp folder

    function __divider(num_rows, [divisor])
        Recursively divides the original Excel file to temp_files until it has no more than
        5000 rows each
        Returns an int, the last row number

@author:
    Venâncio 2000644

@contact:
    venancio.gnr@gmail.com

@organization:
    SDATO - DP - UAF - GNR

@version:
    1.0 (01/09/2014):
        - Implementation of the ExcelSplitter() class.
        - Creation of the functions / class methods:
             __init__(excel_path)
             splitter()
             __create_temp_folder()
             __excel_lines_counter()
             __excel_titles_row()
             __create_temp_excel(start_row, end_row, temp_number, titles_row)
             __divider(num_rows, [divisor])
    
    1.1 (09/12/2014):
        - Added docstrings

    1.2 (28/01/2019):
        - changed MAX_NUM_ROWS from 5000 to 50000

    1.3 (22/11/2021):
        - translated return strings to english
    1.4 (28/11/2021):
        - remove library xlrd and add openpyxl

@since:
    01/09/2014
"""

import os
from openpyxl import load_workbook
import xlwt
from first_row_checker import FirstRowChecker


class ExcelSplitter(object):
    """
    criar lista de paths
    criar pasta Temp
    ler o ficheiro excel
    se tiver menos de 50000 linhas, copia-lo p temp e anexar o path à lista
    se tiver mais, dividir ao meio recursivamente
    """

    def __init__(self, excel_path):
        self.excel_path = os.path.abspath(excel_path)
        self.folder = self.excel_path[:self.excel_path.rfind(os.sep)]
        self.file = self.excel_path[self.excel_path.rfind(os.sep) + 1:]
        self.MAX_NUM_ROWS = 500000

    def __create_temp_folder(self):
        temp_folder = self.folder + os.sep + "Temp"
        if not os.path.exists(temp_folder):
            os.makedirs(temp_folder)

    def __excel_lines_counter(self):
        workbook = load_workbook(self.excel_path)  # xlrd() object
        worksheet_names = [x for x in workbook.get_sheet_names()]  # sheet names
        num_rows = 0

        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)  # sheet.Sheet()
            num_rows += worksheet.max_row - 1  # total num of rows

        return num_rows

    def __excel_titles_row(self):
        workbook = load_workbook(self.excel_path)
        worksheet_names = [x for x in workbook.get_sheet_names()]
        titles_row = []
        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)
            num_rows = worksheet.max_row - 1
            num_cells = worksheet.max_column  # - 1
            curr_row = 0

            while curr_row < num_rows:
                curr_row += 1
                curr_cell = 0
                while curr_cell < num_cells:
                    curr_cell += 1
                    cell_value = worksheet.cell(curr_row, curr_cell).value
                    if curr_row == 1:
                        titles_row.append(cell_value)

        return titles_row

    def __create_temp_excel(self, start_row, end_row, temp_number, titles_row):
        workbook = load_workbook(self.excel_path)  # xlrd() object
        new_workbook = xlwt.Workbook()  # xlwt() object
        worksheet_names = [x for x in workbook.get_sheet_names()]  # sheet names

        # Construction of the worksheets list --------------------------
        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)  # sheet.Sheet()
            new_worksheet = new_workbook.add_sheet(name)
            num_cells = worksheet.max_column  # total num of columns
            curr_row = int(start_row) - 1  # current row

            while curr_row < end_row:  # while curr_row < num_rows:
                curr_row += 1
                curr_cell = 0

                while curr_cell < num_cells:
                    curr_cell += 1
                    if start_row != 0 and curr_row - int(start_row) == 0:
                        cell_value = titles_row[curr_cell]
                    else:
                        cell_value = worksheet.cell_value(curr_row, curr_cell)

                    new_worksheet.write(curr_row - int(start_row), curr_cell,
                                        cell_value)

        new_workbook.save(self.folder + os.sep + 'Temp' + os.sep + 'temp' + str(temp_number) +
                          self.file[:-1])

    def __divider(self, num_rows, divisor=2):
        if num_rows / divisor <= self.MAX_NUM_ROWS:
            return int(num_rows / divisor)
        else:
            divisor += 1
            return self.__divider(num_rows, divisor)

    def splitter(self):
        paths = []

        titles_row = self.__excel_titles_row()

        if not FirstRowChecker(titles_row).check():
            return 'The first 4 columns in Excel must be:\n  \
                    latitude, longitude, name and description'

        if self.__excel_lines_counter() <= self.MAX_NUM_ROWS:
            paths.append(self.folder + os.sep + self.file)
        else:
            self.__create_temp_folder()
            num_rows = self.__excel_lines_counter()
            last_row = self.__divider(num_rows)

            num_of_last_rows = [last_row]
            start_row = [0, last_row]
            next_last_row = 2 * last_row
            start_row.append(next_last_row)

            while next_last_row < num_rows:
                num_of_last_rows.append(next_last_row)
                next_last_row += last_row
                start_row.append(next_last_row)

            if num_rows not in num_of_last_rows:
                num_of_last_rows.append(num_rows)

            for i in range(len(num_of_last_rows)):
                self.__create_temp_excel(start_row[i], num_of_last_rows[i], i,
                                         titles_row)
                temp_file = "temp" + str(i) + self.file[:-1]
                paths.append(self.folder + os.sep + 'Temp' + os.sep + temp_file)

        return paths
