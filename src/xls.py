"""
@python version:
    Python 3.4

@summary:
    Control composed by the XlsControl() class, which reads data from
    an EXEL file and return them in a form of a list.
    The class has two public functions:
        - __init__(file_name),
        - read_exel().

@note:
    function __init__(file_name):
        Class constructor.
        Take as parameters the name of the EXEL file.

    function read_exel():
        Reads the names of the EXEL sheets and append them to a list.
        Then, reads all the cells from that sheets and append their
        values to another list.
        At last, returns a list which combines the sheet names and the
        cell values.

@author:
    Venâncio 2000644

@contact:
    venancio.gnr@gmail.com

@organization:
    SDATO - DP - UAF - GNR

@version:
    1.0 (14/11/2013):
        - Creation of the function abrir_exel(ficheiro)

    1.1 (18/11/2013):
        - Implementation of the XlsControl() class.
        - Update of the abrir_exel() function, which is now named
            read_exel():
            - The EXEL cell values to append to the list, are now
              formated according they are Unicode, ints, floats or
              EXEL datetime values.

    1.2 (06/12/2013):
        - Translation of all comments to english and limitation of the
          maximum line lenght. Following the rules of the PEP 8,
          Style Guide for Python Code, writed by Guido van Rossum,
          Barry Warsaw and Nick Coghlan:
            - "Python coders from non-English speaking countries: please
              write your comments in English, unless you are 120 per
              cent sure that the code will never be read by people who
              don't speak your language."
            - "Limit all lines to a maximum of 79 characters. For
              flowing long blocks of text with fewer structural
              restrictions (docstrings or comments), the line length
              should be limited to 72 characters."
        - Changed the python version to 3.3

    1.3 (09/04/2014):
        - Added a call to the new class FistRowChecker() which breaks the
            program if the first four columns of the Excel file aren't
                - latitude
                - longitude
                - name
                - description

    1.4 (01/12/2014):
        - Added a check for datetime type in excel cell. Get an error if
            the value is superior to 24:00:00

    1.5 (22/11/2021):
        - Remove library xlrd to read Excel and use openpyxl instead

@since:
    14/11/2013
"""
import datetime
from itertools import zip_longest  # to combine lists
from xlrd import xldate_as_tuple  # to read date and time from EXCEL files
from openpyxl import load_workbook
from first_row_checker import FirstRowChecker


class XlsControl(object):
    """
    Reads the data of an EXEL file and returns a list.
    """

    def __init__(self, file_name):
        """
        str -> Xlscontrol() object

        file_name is the name of the EXEL file.
        """
        self.file_name = file_name

    def read_exel(self):
        """
        None -> list

        Reads the names of the EXEL books and append them to a list.
        Then, reads all the cells from that books and append their
        values to another list.
        At last, returns a list which combines the book names and the
        cell values.
        """

        workbook = load_workbook(self.file_name, data_only=True)  # openpyxl object
        worksheet_names = workbook.sheetnames  # sheet names
        worksheets = []  # cell values of all sheets

        # Construction of the worksheets list --------------------------
        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)  # sheet.Sheet()
            num_rows = worksheet.max_row - 1  # total num of rows
            num_cells = worksheet.max_column  # total num of columns
            curr_row = 0  # current row
            data_rows = []  # data from all rows

            while curr_row < num_rows:
                data_row = []  # data from the current row
                curr_row += 1
                curr_cell = 0

                while curr_cell < num_cells:
                    curr_cell += 1
                    cell_value = worksheet.cell(curr_row, curr_cell).value
                    # Formation of the cell values to append: --------
                    # in Exel all ints are floats terminated in .0 -------
                    if type(cell_value) is float and str(
                            cell_value)[-2:] == ".0" and \
                            worksheet.cell_type(curr_row, curr_cell) is int:
                        data_row.append(int(cell_value))
                    # datetime -----------------------------------------
                    elif type(cell_value) is tuple or type(cell_value) is datetime.date:
                        #  elif worksheet.cell_type(curr_row, curr_cell) is datetime.date:
                        try:
                            xldate_as_tuple(abs(cell_value), 0)
                            data_row.append(
                                xldate_as_tuple(abs(cell_value), 0))
                        except:
                            total_secs = int(cell_value * 24 * 3600)
                            hours = total_secs // 3600
                            mins = (total_secs % 3600) // 60
                            secs = total_secs % 60
                            duration = str(hours) + ':' + str(mins) + ':' + str(secs)
                            data_row.append(duration)

                            # logfile = open('error.log', 'a')

                    elif type(cell_value) is float:  # real floats -----
                        data_row.append(cell_value)
                    else:  # unicode -----------------------------------
                        try:  # In some files all values are unicode, so
                            if str(cell_value[-2] == ".0"):
                                data_row.append(int(cell_value))
                            else:
                                data_row.append(float(cell_value))
                        except:
                            data_row.append(cell_value)

                data_rows.append(data_row)

            if not FirstRowChecker(data_rows[0]).check():
                return 'The first 4 columns of Excel file must be:\n\
                latitude, longitude, name and description.'

            worksheets.append(data_rows)

        # 3D list combined of the sheet names and the sheet cell values:
        # [[[1st sheet name][1st row data][2nd row data][...][Nth row data]]
        # [[2nd sheet name][1st row data][...]]...]
        return [[[x]] + y for x, y in zip_longest(worksheet_names, worksheets)]

