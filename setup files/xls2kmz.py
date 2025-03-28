"""
@python version:
    Python 3.4

@summary:
    Copy of all interface and controls to use with setup.py so it
        builds the programm

@note:
    See all particular modules

@author:
    Venâncio 2000644

@contact:
    venancio.gnr@gmail.com

@organization:
    SDATO - DP - UAF - GNR

@version:
    See all particular modules

@since:
    15/11/2013
"""
import datetime
import logging  # to create a logfile
import math  # calculate the inclination of the arows
# ----------------------------------------------------------------------
import os  # to manipulate paths
import random  # to generate random numbers
import sys  # to redirect sys.stderr to a log file
import threading  # to use varios threads at same time
import xml.etree.ElementTree as et  # build mkd files
import zipfile  # to zip and unzip files, rebuild the kmz
from itertools import zip_longest  # to combine lists
from time import sleep  # delay events in the window
# for the windowed interface -------------------------------------------
from tkinter import BOTTOM, CENTER, FALSE, E, W, EW, SUNKEN, TOP, TRUE, BOTH, \
    LEFT, Tk, Frame, Label, Button, Message, Menu, Toplevel, StringVar, \
    BooleanVar, filedialog, ttk

import simplekml  # build KML/KMZ files
from openpyxl import load_workbook # to read EXCEL files
from xlrd  import xldate_as_tuple # to read EXCEL dates
import xlwt  # to write EXCEL files
from PIL import ImageTk, Image


class FirstRowChecker(object):
    """
    Verifica se as primeiras 4 colunas do Excel
    são: latitude, longitude, name e description
    """

    def __init__(self, first_row):
        """
        first_row é uma lista
        """
        self.first_row = first_row

    def check(self):
        low = [x.lower() for x in self.first_row]
        eng = ['latitude', 'longitude', 'name', 'description']
        por = ['latitude', 'longitude', 'nome', 'descricao']
        ptr = ['latitude', 'longitude', 'nome', 'descrição']

        if low[0:4] == eng or low[0:4] == por or low[0:4] == ptr:
            return True
        else:
            return False


class ExcelSplitter(object):
    """
    criar lista de paths
    criar pasta Temp
    ler o ficheiro excel
    se tiver menos de 50000 linhas, copia-lo p temp e anexar o path à lista
    se tiver mais, dividir ao meio recursivamente
    """

    def __init__(self, excel_path):
        self.excel_path = os.path.abspath(excel_path)
        self.folder = self.excel_path[:self.excel_path.rfind(os.sep)]
        self.file = self.excel_path[self.excel_path.rfind(os.sep) + 1:]
        self.MAX_NUM_ROWS = 500000

    def __create_temp_folder(self):
        temp_folder = self.folder + os.sep + "Temp"
        if not os.path.exists(temp_folder):
            os.makedirs(temp_folder)

    def __excel_lines_counter(self):
        workbook = load_workbook(self.excel_path)  # xlrd() object
        worksheet_names = [x for x in workbook.get_sheet_names()]  # sheet names
        num_rows = 0

        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)  # sheet.Sheet()
            num_rows += worksheet.max_row - 1  # total num of rows

        return num_rows

    def __excel_titles_row(self):
        workbook = load_workbook(self.excel_path)
        worksheet_names = [x for x in workbook.get_sheet_names()]
        titles_row = []
        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)
            num_rows = worksheet.max_row - 1
            num_cells = worksheet.max_column  # - 1
            curr_row = 0

            while curr_row < num_rows:
                curr_row += 1
                curr_cell = 0
                while curr_cell < num_cells:
                    curr_cell += 1
                    cell_value = worksheet.cell(curr_row, curr_cell).value
                    if curr_row == 1:
                        titles_row.append(cell_value)

        return titles_row

    def __create_temp_excel(self, start_row, end_row, temp_number, titles_row):
        workbook = load_workbook(self.excel_path)  # xlrd() object
        new_workbook = xlwt.Workbook()  # xlwt() object
        worksheet_names = [x for x in workbook.get_sheet_names()]  # sheet names

        # Construction of the worksheets list --------------------------
        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)  # sheet.Sheet()
            new_worksheet = new_workbook.add_sheet(name)
            num_cells = worksheet.max_column  # total num of columns
            curr_row = int(start_row) - 1  # current row

            while curr_row < end_row:  # while curr_row < num_rows:
                curr_row += 1
                curr_cell = 0

                while curr_cell < num_cells:
                    curr_cell += 1
                    if start_row != 0 and curr_row - int(start_row) == 0:
                        cell_value = titles_row[curr_cell]
                    else:
                        cell_value = worksheet.cell_value(curr_row, curr_cell)

                    new_worksheet.write(curr_row - int(start_row), curr_cell,
                                        cell_value)

        new_workbook.save(self.folder + os.sep + 'Temp' + os.sep + 'temp' + str(temp_number) +
                          self.file[:-1])

    def __divider(self, num_rows, divisor=2):
        if num_rows / divisor <= self.MAX_NUM_ROWS:
            return int(num_rows / divisor)
        else:
            divisor += 1
            return self.__divider(num_rows, divisor)

    def splitter(self):
        paths = []

        titles_row = self.__excel_titles_row()

        if not FirstRowChecker(titles_row).check():
            return 'The first 4 columns in Excel must be:\n  \
                    latitude, longitude, name and description'

        if self.__excel_lines_counter() <= self.MAX_NUM_ROWS:
            paths.append(self.folder + os.sep + self.file)
        else:
            self.__create_temp_folder()
            num_rows = self.__excel_lines_counter()
            last_row = self.__divider(num_rows)

            num_of_last_rows = [last_row]
            start_row = [0, last_row]
            next_last_row = 2 * last_row
            start_row.append(next_last_row)

            while next_last_row < num_rows:
                num_of_last_rows.append(next_last_row)
                next_last_row += last_row
                start_row.append(next_last_row)

            if num_rows not in num_of_last_rows:
                num_of_last_rows.append(num_rows)

            for i in range(len(num_of_last_rows)):
                self.__create_temp_excel(start_row[i], num_of_last_rows[i], i,
                                         titles_row)
                temp_file = "temp" + str(i) + self.file[:-1]
                paths.append(self.folder + os.sep + 'Temp' + os.sep + temp_file)

        return paths


class KmlJoiner(object):

    def __init__(self, list_of_kmls, original_path):
        self.kmls = list_of_kmls
        self.path = original_path

    def __extract_all(self):
        """
        ZipFile.extractall([path[, members[, pwd]]])

        Extract all members from the archive to the current working directory.
        path specifies a different directory to extract to.
        members must be a subset from list returned by namelist().
        pwd is the password used for encrypted files.
        """
        i = 0

        for k in self.kmls:
            zf = zipfile.ZipFile(k, 'r')
            zf.extractall()
            if 'doc.kml' in os.listdir():
                os.rename('doc.kml', 'doc' + str(i) + '.kml')
                i += 1
            zf.close()  # acho que não é preciso

    def build_new_kmz(self):
        # change working directory p "Temp"
        os.chdir(os.path.dirname(self.kmls[0]))
        self.__extract_all()

        zf = zipfile.ZipFile(self.path, "a")  # make new kmz in parent dir
        doc_list = [doc for doc in os.listdir() if doc[-3:] == 'kml']
        self.__join_docs(doc_list)
        zf.write('doc.kml')

        os.chdir(os.path.dirname(self.kmls[0]) + '\\files')
        for image in os.listdir():
            zf.write(image, arcname='files\\' + image)

        zf.close()

    def __join_docs(self, doc_list):
        os.system("copy *.kml doc_temp.kml")
        self.__kml_parser('doc_temp.kml')

    def __kml_parser(self, doc):
        close_folder = False
        open_folder = True

        lines = open(doc, 'r', encoding='utf-8').readlines()  # list of lines
        new_kml = open('doc.kml', 'w')

        for line in lines[:-4]:
            if '</Folder>' in line:
                close_folder = True
                open_folder = False
            if close_folder is False and open_folder is True:
                new_kml.writelines(line)
            if '<Folder' in line:
                close_folder = False
                open_folder = True
        new_kml.writelines(lines[-4:-1])

        new_kml.close()


class TempCleaner(object):

    def __init__(self, temp_folder):
        self.folder = temp_folder

    def clean(self):
        try:
            os.chdir(self.folder)
            self.__sub_folder_crawler()
            self.__file_deleter()
            os.chdir(os.pardir)
            self.__folder_deleter(self.folder)
        except OSError:
            self.clean()

    def __file_deleter(self):
        for files in os.listdir():
            os.remove(files)

    def __folder_deleter(self, directory):
        os.rmdir(directory)

    def __sub_folder_searcher(self):
        return [f for f in os.listdir() if os.path.isdir(f)]

    def __sub_folder_crawler(self):
        sub_folders = self.__sub_folder_searcher()
        if len(sub_folders) > 0:
            for f in sub_folders:
                os.chdir(f)
                if len(os.listdir()) == 0:
                    os.chdir(os.pardir)
                    self.__folder_deleter(f)
                else:
                    return self.__sub_folder_crawler()
        else:
            return os.getcwd()


class LogFile(object):
    """
    File-like object to log text using the `logging` module.

    http://stackoverflow.com/questions/616645/how-do-i-duplicate-
    sys-stdout-to-a-log-file-in-python/3423392#3423392
    """

    def __init__(self, name=None):
        self.logger = logging.getLogger(name)

    def write(self, msg, level=logging.INFO):
        self.logger.log(level, msg)

    def flush(self):
        for handler in self.logger.handlers:
            handler.flush()


class XlsControl(object):
    """
    Reads the data of an EXEL file and returns a list.
    """

    def __init__(self, file_name):
        """
        str -> Xlscontrol() object

        file_name is the name of the EXEL file.
        """
        self.file_name = file_name

    def read_exel(self):
        """
        None -> list

        Reads the names of the EXEL books and append them to a list.
        Then, reads all the cells from that books and append their
        values to another list.
        At last, returns a list which combines the book names and the
        cell values.
        """

        workbook = load_workbook(self.file_name, data_only=True)  # openpyxl object
        worksheet_names = workbook.sheetnames  # sheet names
        worksheets = []  # cell values of all sheets

        # Construction of the worksheets list --------------------------
        for name in worksheet_names:
            worksheet = workbook.get_sheet_by_name(name)  # sheet.Sheet()
            num_rows = worksheet.max_row - 1  # total num of rows
            num_cells = worksheet.max_column  # total num of columns
            curr_row = 0  # current row
            data_rows = []  # data from all rows

            while curr_row < num_rows:
                data_row = []  # data from the current row
                curr_row += 1
                curr_cell = 0

                while curr_cell < num_cells:
                    curr_cell += 1
                    cell_value = worksheet.cell(curr_row, curr_cell).value
                    # Formation of the cell values to append: --------
                    # in Exel all ints are floats terminated in .0 -------
                    if type(cell_value) is float and str(
                            cell_value)[-2:] == ".0" and \
                            worksheet.cell_type(curr_row, curr_cell) is int:
                        data_row.append(int(cell_value))
                    # datetime -----------------------------------------
                    elif type(cell_value) is tuple or type(cell_value) is datetime.date:
                        #  elif worksheet.cell_type(curr_row, curr_cell) is datetime.date:
                        try:
                            xldate_as_tuple(abs(cell_value), 0)
                            data_row.append(
                                xldate_as_tuple(abs(cell_value), 0))
                        except:
                            total_secs = int(cell_value * 24 * 3600)
                            hours = total_secs // 3600
                            mins = (total_secs % 3600) // 60
                            secs = total_secs % 60
                            duration = str(hours) + ':' + str(mins) + ':' + str(secs)
                            data_row.append(duration)

                            # logfile = open('error.log', 'a')

                    elif type(cell_value) is float:  # real floats -----
                        data_row.append(cell_value)
                    else:  # unicode -----------------------------------
                        try:  # In some files all values are unicode, so
                            if str(cell_value[-2] == ".0"):
                                data_row.append(int(cell_value))
                            else:
                                data_row.append(float(cell_value))
                        except:
                            data_row.append(cell_value)

                data_rows.append(data_row)

            if not FirstRowChecker(data_rows[0]).check():
                return 'The first 4 columns of Excel file must be:\n\
                latitude, longitude, name and description.'

            worksheets.append(data_rows)

        # 3D list combined of the sheet names and the sheet cell values:
        # [[[1st sheet name][1st row data][2nd row data][...][Nth row data]]
        # [[2nd sheet name][1st row data][...]]...]
        return [[[x]] + y for x, y in zip_longest(worksheet_names, worksheets)]


class MyThread(threading.Thread):
    """Subclass Thread and Create subclass instance"""

    def __init__(self, function, args, name=''):
        """
        Take as parameters "function", a string, the name of the
        function to run as a thread, "args", a tuple of strings
        with the function arguments, and optionally a "name" for
        the tread.
        """
        threading.Thread.__init__(self)
        self.name = name
        self.func = function
        self.args = args
        self.res = None

    def get_result(self):
        """Retrieves the value of self.res"""
        return self.res

    def run(self):
        """Applies the arguments to the function running as thread"""
        self.res = self.func(*self.args)


class ColornameToKml(object):
    """Translates color names into hexadecimal values."""

    def __init__(self, colorname):
        """
        str -> () simplekml.Kml.color() object

        Builds a dictionary where the key is a color name and the value
        is a simplekml.Kml.color() object of that color.
        """
        self.colorname = colorname.lower().strip()
        color = simplekml.Color
        self.colorname_to_kml = {'aliceblue': color.aliceblue,
                                 'antiquewhite': color.antiquewhite,
                                 'aqua': color.aqua,
                                 'aquamarine': color.aquamarine,
                                 'azure': color.azure,
                                 'beige': color.beige,
                                 'bisque': color.bisque,
                                 'black': color.black,
                                 'blanchedalmond ': color.blanchedalmond,
                                 'blue': color.blue,
                                 'blueviolet': color.blueviolet,
                                 'brown': color.brown,
                                 'burlywood': color.burlywood,
                                 'cadetblue': color.cadetblue,
                                 'chartreuse': color.chartreuse,
                                 'chocolate': color.chocolate,
                                 'coral': color.coral,
                                 'cornflowerblue': color.cornflowerblue,
                                 'cornsilk': color.cornsilk,
                                 'crimson': color.crimson,
                                 'cyan': color.cyan,
                                 'darkblue': color.darkblue,
                                 'darkcyan': color.darkcyan,
                                 'darkgoldenrod': color.darkgoldenrod,
                                 'darkgrey': color.darkgrey,
                                 'darkgreen': color.darkgreen,
                                 'darkkhaki': color.darkkhaki,
                                 'darkmagenta': color.darkmagenta,
                                 'darkolivegreen': color.darkolivegreen,
                                 'darkorange': color.darkorange,
                                 'darkorchid': color.darkorchid,
                                 'darkred': color.darkred,
                                 'darksalmon': color.darksalmon,
                                 'darkseagreen': color.darkseagreen,
                                 'darkslateblue': color.darkslateblue,
                                 'darkslategray': color.darkslategray,
                                 'darkturquoise': color.darkturquoise,
                                 'darkviolet': color.darkviolet,
                                 'deeppink': color.deeppink,
                                 'deepskyblue': color.deepskyblue,
                                 'dimgray': color.dimgray,
                                 'dodgerblue': color.dodgerblue,
                                 'firebrick': color.firebrick,
                                 'floralwhite': color.floralwhite,
                                 'forestgreen': color.forestgreen,
                                 'fuchsia': color.fuchsia,
                                 'gainsboro': color.gainsboro,
                                 'ghostwhite': color.ghostwhite,
                                 'gold': color.gold,
                                 'goldenrod': color.goldenrod,
                                 'gray': color.gray,
                                 'green': color.green,
                                 'greenyellow': color.greenyellow,
                                 'honeydew': color.honeydew,
                                 'hotpink': color.hotpink,
                                 'indianred': color.indianred,
                                 'indigo': color.indigo,
                                 'ivory': color.ivory,
                                 'khaki': color.khaki,
                                 'lavender': color.lavender,
                                 'lavenderblush': color.lavenderblush,
                                 'lawngreen': color.lawngreen,
                                 'lemonchiffon': color.lemonchiffon,
                                 'lightblue': color.lightblue,
                                 'lightcoral': color.lightcoral,
                                 'lightcyan': color.lightcyan,
                                 'lightgoldenrodyellow':
                                     color.lightgoldenrodyellow,
                                 'lightgray': color.lightgray,
                                 'lightgreen': color.lightgreen,
                                 'lightpink': color.lightpink,
                                 'lightsalmon': color.lightsalmon,
                                 'lightseagreen': color.lightseagreen,
                                 'lightskyblue': color.lightskyblue,
                                 'lightslategray': color.lightslategray,
                                 'lightsteelblue': color.lightsteelblue,
                                 'lightyellow': color.lightyellow,
                                 'lime': color.lime,
                                 'limegreen': color.limegreen,
                                 'linen': color.linen,
                                 'magenta': color.magenta,
                                 'maroon': color.maroon,
                                 'mediumaquamarine': color.mediumaquamarine,
                                 'mediumblue': color.mediumblue,
                                 'mediumorchid': color.mediumorchid,
                                 'mediumpurple': color.mediumpurple,
                                 'mediumseagreen': color.mediumseagreen,
                                 'mediumslateblue': color.mediumslateblue,
                                 'mediumspringgreen': color.mediumspringgreen,
                                 'mediumturquoise': color.mediumturquoise,
                                 'mediumvioletred': color.mediumvioletred,
                                 'midnightblue': color.midnightblue,
                                 'mintcream': color.mintcream,
                                 'mistyrose': color.mistyrose,
                                 'moccasin': color.moccasin,
                                 'navajowhite': color.navajowhite,
                                 'navy': color.navy,
                                 'oldlace': color.oldlace,
                                 'olive': color.olive,
                                 'olivedrab': color.olivedrab,
                                 'orange': color.orange,
                                 'orangered': color.orangered,
                                 'orchid': color.orchid,
                                 'palegoldenrod': color.palegoldenrod,
                                 'palegreen': color.palegreen,
                                 'paleturquoise': color.paleturquoise,
                                 'palevioletred': color.palevioletred,
                                 'papayawhip': color.papayawhip,
                                 'peachpuff': color.peachpuff,
                                 'peru': color.peru,
                                 'pink': color.pink,
                                 'plum': color.plum,
                                 'powderblue': color.powderblue,
                                 'purple': color.purple,
                                 'red': color.red,
                                 'rosybrown': color.rosybrown,
                                 'royalblue': color.royalblue,
                                 'saddlebrown': color.saddlebrown,
                                 'salmon': color.salmon,
                                 'sandybrown': color.sandybrown,
                                 'seagreen': color.seagreen,
                                 'seashell': color.seashell,
                                 'sienna': color.sienna,
                                 'silver': color.silver,
                                 'skyblue': color.skyblue,
                                 'slateblue': color.slateblue,
                                 'slategray': color.slategray,
                                 'snow': color.snow,
                                 'springgreen': color.springgreen,
                                 'steelblue': color.steelblue,
                                 'tan': color.tan,
                                 'teal': color.teal,
                                 'thistle': color.thistle,
                                 'tomato': color.tomato,
                                 'turquoise': color.turquoise,
                                 'violet': color.violet,
                                 'wheat': color.wheat,
                                 'white': color.white,
                                 'whitesmoke': color.whitesmoke,
                                 'yellow': color.yellow,
                                 'yellowgreen': color.yellowgreen}

    def get_color(self):
        return self.colorname_to_kml[self.colorname]


class ColornameToHex(object):
    """Translates color names into hexadecimal values."""

    def __init__(self, colorname):
        """
        str -> Colorname() object

        Builds a dictionary where the key is a color name and the value
        is the hexadecimal value in RGB.
        """
        self.colorname = colorname.lower().strip()
        self.colorname_to_hex = {'aliceblue': '#F0F8FF',
                                 'antiquewhite': '#FAEBD7',
                                 'aqua': '#00FFFF',
                                 'aquamarine': '#7FFFD4',
                                 'azure': '#F0FFFF',
                                 'beige': '#F5F5DC',
                                 'bisque': '#FFE4C4',
                                 'black': '#000000',
                                 'blanchedalmond': '#FFEBCD',
                                 'blue': '#0000FF',
                                 'blueviolet': '#8A2BE2',
                                 'brown': '#A52A2A',
                                 'burlywood': '#DEB887',
                                 'cadetblue': '#5F9EA0',
                                 'chartreuse': '#7FFF00',
                                 'chocolate': '#D2691E',
                                 'coral': '#FF7F50',
                                 'cornflowerblue': '#6495ED',
                                 'cornsilk': '#FFF8DC',
                                 'crimson': '#DC143C',
                                 'cyan': '#00FFFF',
                                 'darkblue': '#00008B',
                                 'darkcyan': '#008B8B',
                                 'darkgoldenrod': '#B8860B',
                                 'darkgrey': '#A9A9A9',
                                 'darkgreen': '#006400',
                                 'darkkhaki': '#BDB76B',
                                 'darkmagenta': '#8B008B',
                                 'darkolivegreen': '#556B2F',
                                 'darkorange': '#FF8C00',
                                 'darkorchid': '#9932CC',
                                 'darkred': '#8B0000',
                                 'darksalmon': '#E9967A',
                                 'darkseagreen': '#8FBC8F',
                                 'darkslateblue': '#483D8B',
                                 'darkslategray': '#2F4F4F',
                                 'darkturquoise': '#00CED1',
                                 'darkviolet': '#9400D3',
                                 'deeppink': '#FF1493',
                                 'deepskyblue': '#00BFFF',
                                 'dimgray': '#696969',
                                 'dodgerblue': '#1E90FF',
                                 'firebrick': '#B22222',
                                 'floralwhite': '#FFFAF0',
                                 'forestgreen': '#228B22',
                                 'fuchsia': '#FF00FF',
                                 'gainsboro': '#DCDCDC',
                                 'ghostwhite': '#F8F8FF',
                                 'gold': '#FFD700',
                                 'goldenrod': '#DAA520',
                                 'gray': '#808080',
                                 'green': '#008000',
                                 'greenyellow': '#ADFF2F',
                                 'honeydew': '#F0FFF0',
                                 'hotpink': '#FF69B4',
                                 'indianred': '#CD5C5C',
                                 'indigo': '#4B0082',
                                 'ivory': '#FFFFF0',
                                 'khaki': '#F0E68C',
                                 'lavender': '#E6E6FA',
                                 'lavenderblush': '#FFF0F5',
                                 'lawngreen': '#7CFC00',
                                 'lemonchiffon': '#FFFACD',
                                 'lightblue': '#ADD8E6',
                                 'lightcoral': '#F08080',
                                 'lightcyan': '#E0FFFF',
                                 'lightgoldenrodyellow': '#FAFAD2',
                                 'lightgray': '#D3D3D3',
                                 'lightgreen': '#90EE90',
                                 'lightpink': '#FFB6C1',
                                 'lightsalmon': '#FFA07A',
                                 'lightseagreen': '#20B2AA',
                                 'lightskyblue': '#87CEFA',
                                 'lightslategray': '#778899',
                                 'lightsteelblue': '#B0C4DE',
                                 'lightyellow': '#FFFFE0',
                                 'lime': '#00FF00',
                                 'limegreen': '#32CD32',
                                 'linen': '#FAF0E6',
                                 'magenta': '#FF00FF',
                                 'maroon': '#800000',
                                 'mediumaquamarine': '#66CDAA',
                                 'mediumblue': '#0000CD',
                                 'mediumorchid': '#BA55D3',
                                 'mediumpurple': '#9370DB',
                                 'mediumseagreen': '#3CB371',
                                 'mediumslateblue': '#7B68EE',
                                 'mediumspringgreen': '#00FA9A',
                                 'mediumturquoise': '#48D1CC',
                                 'mediumvioletred': '#C71585',
                                 'midnightblue': '#191970',
                                 'mintcream': '#F5FFFA',
                                 'mistyrose': '#FFE4E1',
                                 'moccasin': '#FFE4B5',
                                 'navajowhite': '#FFDEAD',
                                 'navy': '#000080',
                                 'oldlace': '#FDF5E6',
                                 'olive': '#808000',
                                 'olivedrab': '#6B8E23',
                                 'orange': '#FFA500',
                                 'orangered': '#FF4500',
                                 'orchid': '#DA70D6',
                                 'palegoldenrod': '#EEE8AA',
                                 'palegreen': '#98FB98',
                                 'paleturquoise': '#AFEEEE',
                                 'palevioletred': '#DB7093',
                                 'papayawhip': '#FFEFD5',
                                 'peachpuff': '#FFDAB9',
                                 'peru': '#CD853F',
                                 'pink': '#FFC0CB',
                                 'plum': '#DDA0DD',
                                 'powderblue': '#B0E0E6',
                                 'purple': '#800080',
                                 'red': '#FF0000',
                                 'rosybrown': '#BC8F8F',
                                 'royalblue': '#4169E1',
                                 'saddlebrown': '#8B4513',
                                 'salmon': '#FA8072',
                                 'sandybrown': '#F4A460',
                                 'seagreen': '#2E8B57',
                                 'seashell': '#FFF5EE',
                                 'sienna': '#A0522D',
                                 'silver': '#C0C0C0',
                                 'skyblue': '#87CEEB',
                                 'slateblue': '#6A5ACD',
                                 'slategray': '#708090',
                                 'snow': '#FFFAFA',
                                 'springgreen': '#00FF7F',
                                 'steelblue': '#4682B4',
                                 'tan': '#D2B48C',
                                 'teal': '#008080',
                                 'thistle': '#D8BFD8',
                                 'tomato': '#FF6347',
                                 'turquoise': '#40E0D0',
                                 'violet': '#EE82EE',
                                 'wheat': '#F5DEB3',
                                 'white': '#FFFFFF',
                                 'whitesmoke': '#F5F5F5',
                                 'yellow': '#FFFF00',
                                 'yellowgreen': '#9ACD32'}

    def get_rgb(self):
        """
        None -> str

        Returns a hexadecimal RGB value.
        """
        return self.colorname_to_hex[self.colorname].lower()

    def get_rgba(self):
        """
        None -> str

        Returns a hexadecimal RGBA value.
        """
        return self.colorname_to_hex[self.colorname].lower() + 'ff'

    def get_bgr(self):
        """
        None -> str

        Returns a hexadecimal BGR value.
        """
        b = self.colorname_to_hex[self.colorname][-2:].lower()
        g = self.colorname_to_hex[self.colorname][3:5].lower()
        r = self.colorname_to_hex[self.colorname][1:3].lower()

        return "#" + b + g + r

    def get_abgr(self):
        """
        None -> str

        Returns a hexadecimal ABGR value.
        """
        bgr = self.get_bgr()
        return bgr[0] + "ff" + bgr[1:]


class Coordinates(object):
    """
    Creates a coordinate object with latitude and longitude and convert
    it to the Decimal Degrees format.

    The types of coordinates formats are:
    Decimal Degrees: 32.8303ºN ; 116.7762ºW
    Degrees Minutes Seconds:32º49'49''N ; 116º46'34''W
    Hybrid: 32º49.818'N ; 116º46.574'W
    """

    def __init__(self, latitude, longitude):
        """
        str, str -> None

        Class constructor.
        Take as parameters "latitude" and "longitude", two strings.
        Builds two class variables lat and lon.
        First removes any white spaces from latitude and longitude,
        calling the __remove_spaces()
        then removes the cardinal direction letter calling __remove_chars.
        """
        self.__lat = self.__remove_spaces(latitude)
        self.__lon = self.__remove_spaces(longitude)
        self.lat = self.__remove_chars(self.__lat)
        self.lon = self.__remove_chars(self.__lon)

    def convert(self):
        """
        None -> tuple

        Checks for the format of the coordinates and call __convert_dms,
        if the format is degrees, minutes, seconds; call __convert_h,
        if the format is hybrid aka degrees, decimal minutes.
        returns lat and lon in decimal degrees format
        """
        coords_format = self.__get_format(self.lat)

        if coords_format == "dms":
            self.__convert_dms()
        elif coords_format == "h":
            self.__convert_h()

        return self.lat, self.lon

    def __remove_spaces(self, string):
        """
        str -> str

        Returns the string whithout spaces or tabs
        """
        string = string.replace("\t", "")
        return string.replace(" ", "")

    def __remove_chars(self, string):
        """
        str -> str

        Returns the string whithout the cardinal direction letter.
        If the cardinal direction letter is 'S' or 'W' it appends a '-'
        (minus sigh) to the begining of the string.
        """
        pos_chars = ['N', 'E', 'n', 'e']
        neg_chars = ['S', 'W', 's', 'w']
        new_string = ''

        for char in string:
            if char in pos_chars:
                new_string += ""
            elif char in neg_chars:
                new_string = "-" + new_string
                "-" + string.replace(char, "")
            else:
                new_string += char

        return new_string

    def __get_format(self, string):
        """
        str -> str

        Check the format of the string and returns "dms", "h" or "dd",
        acording to Degrees Minutes Seconds format, Hybrid format, or
        Decimal Degrees format, respectivaly
        """
        m = "'"
        s = "''"
        s2 = '"'
        s3 = "º"
        s4 = "°"
        s5 = ","
        s6 = "."

        if (s in string or s2 in string) or \
                (s3 in string and not s5 in string) or \
                (s3 in string and not s6 in string) or \
                (s4 in string and not s5 in string) or \
                (s4 in string and not s6 in string):
            return 'dms'
        if m in string and s not in string:
            return 'h'
        else:
            return 'dd'

    def __convert_dms(self):
        """
        None -> None

        Convert lat and lon from Degrees Minutes Seconds format to
        Decimal Degrees format
        Decimal Degrees = Degrees + Minutes / 60 + Seconds / 3600
        """
        lat = self.lat.replace('"', "''")
        lon = self.lon.replace('"', "''")
        lat_sign = ""
        lon_sign = ""

        if "º" in lat:
            if "-0" in lat[:lat.index("º")]:
                lat_sign = "-"
            else:
                pass
            lat_deg = int(lat[:lat.index("º")])
            if "-0" in lon[:lon.index("º")]:
                lon_sign = "-"
            else:
                pass
            lon_deg = int(lon[:lon.index("º")])
            lat_min = int(lat[lat.index("º") + 1:lat.index("'")])
            lon_min = int(lon[lon.index("º") + 1:lon.index("'")])
        else:
            if "-0" in lat[:lat.index("°")]:
                lat_sign = "-"
                lat_deg = 0
            else:
                lat_deg = int(lat[:lat.index("°")])

            if "-0" in lon[:lon.index("°")]:
                lon_sign = "-"
                lon_deg = 0
            else:
                lon_deg = int(lon[:lon.index("°")])

            lat_min = int(lat[lat.index("°") + 1:lat.index("'")])
            lon_min = int(lon[lon.index("°") + 1:lon.index("'")])

        if "''" in lat:
            lat_sec = float(lat[lat.index("'") + 1:lat.index("''")])
        else:
            lat_sec = float(lat[lat.index("'") + 1:])
        if "''" in lon:
            lon_sec = float(lon[lon.index("'") + 1:lon.index("''")])
        else:
            lon_sec = float(lon[lon.index("'") + 1:])

        if lat_deg < 0:
            lat_sign = "-"
        if lon_deg < 0:
            lon_sign = "-"

        int_lat = abs(lat_deg) + (lat_min / 60.0) + (lat_sec / 3600.0)
        int_lon = abs(lon_deg) + (lon_min / 60.0) + (lon_sec / 3600.0)
        self.lat = lat_sign + str(int_lat)
        self.lon = lon_sign + str(int_lon)

        self.convert()

    def __convert_h(self):
        """
        None -> None

        Convert lat and lon from Hybrid format to Decimal Degrees format
        Decimal Degrees = Degrees + Minutes / 60
        """
        lat = self.lat
        lon = self.lon
        lat_sign = ""
        lon_sign = ""

        try:
            lat_deg = int(lat[:lat.index("º")])
            lat_min = int(lat[lat.index("º") + 1:lat.index("'")])

            lon_deg = float(lon[:lon.index("º")])
            lon_min = float(lon[lon.index("º") + 1:lon.index("'")])
        except:
            lat_deg = int(lat[:lat.index("°")])
            lat_min = int(lat[lat.index("°") + 1:lat.index("'")])

            lon_deg = float(lon[:lon.index("°")])
            lon_min = float(lon[lon.index("°") + 1:lon.index("'")])

        if lat_deg < 0:
            lat_sign = "-"
        if lon_deg < 0:
            lon_sign = "-"

        int_lat = abs(lat_deg) + (lat_min / 60.0)
        int_lon = abs(lon_deg) + (lon_min / 60.0)
        self.lat = lat_sign + str(int_lat)
        self.lon = lon_sign + str(int_lon)

        self.convert()


class KmlControl(object):
    """
    Turns a list into a Kml() object and saves in the disk as a KMZ file.
    """

    def __init__(self, data_list, file_name):
        """
        list, string -> KmlControl() object

        data_list is the EXEL data list.
        file_name is the file name of the KMZ we want to build.
        """
        self.data_list = data_list
        self.file_name = file_name
        self.images_list = []  # list of all icons, photos and images used

    def build_kml(self):
        """
        None -> Kml() object

        Builds a Kml() object from a list passed by the class constructor
        and returns it.

        See: http://simplekml.readthedocs.org/en/latest/index.html
        """
        kml = simplekml.Kml()
        icons = os.getcwd() + os.sep + "icons" + os.sep  # path
        add = "appenddatacolumnstodescription"
        # Build a legend if exist ----------------------------------
        try:
            path = os.getcwd() + os.sep + "fotos" + os.sep + "legenda.png"
            screen = kml.newscreenoverlay(name="Legenda")
            screen.icon.href = path
            screen.overlayxy = simplekml.OverlayXY(x=0, y=1,
                                                   xunits=simplekml.Units.fraction,
                                                   yunits=simplekml.Units.fraction)
            screen.screenxy = simplekml.ScreenXY(x=15, y=15,
                                                 xunits=simplekml.Units.pixels,
                                                 yunits=simplekml.Units.insetpixels)
            screen.size.x = -1
            screen.size.y = -1
            screen.size.xunits = simplekml.Units.fraction
            screen.size.yunits = simplekml.Units.fraction
        except:
            pass

        for item in self.data_list:
            if len(item) > 1:
                folder = kml.newfolder(name=str(item[0][0]))  # sheetname
                headers = [x.lower() for x in item[1]]  # column titles
                # Optional column names to build a point-----------------
                col_names = ["icon", "iconcolor", "iconscale", "description",
                             "appenddatacolumnstodescription", "iconheading",
                             "linestringcolor", "foto", "polygon",
                             "polygoncolor", "polygonaltitude",
                             "polygonazimute", "polygonamplitude",
                             "squarealtitude", "squarelatitude",
                             "squarelongitude", "squarecolor"]
                # See what optional names are in the EXEL data ----------
                next_coords = None
                line = None
                line_color = None
                col_as_name = [True if x in headers else False for x in col_names]
                #  POINT BUILD ##########################################
                for i in range(2, len(item)):
                    # Coordinates ---------------------------------------
                    lat1 = str(item[i][0])
                    lon1 = str(item[i][1])
                    coords = [Coordinates(lon1, lat1).convert()]
                    if i < len(item) - 1:  # Next coordinate (icon heading)
                        lat2 = str(item[i + 1][0])
                        lon2 = str(item[i + 1][1])
                        next_coords = [Coordinates(lon2, lat2).convert()]
                    #  Point --------------------------------------------
                    try:
                        name = self.__formated_time(item[i][2])
                    except:
                        try:
                            name = str(item[i][2])
                        except ValueError:
                            name = item[i][2]
                    #  Icon ---------------------------------------------
                    if col_as_name[col_names.index("icon")] and item[i][headers.index("icon")] != '':
                        point = folder.newpoint(name=name, coords=coords)
                        point.lookat.longitude = coords[0][0]
                        point.lookat.latitude = coords[0][1]
                        point.lookat.altitude = 0
                        point.lookat.heading = 0
                        point.lookat.tilt = 0
                        point.lookat.range = 1230
                        url = icons + str(int(item[i][headers.index("icon")])) + ".png"
                        point.style.iconstyle.icon.href = url
                        point.style.balloonstyle.text = self.__point_description(headers, [item[i][headers.index(add)]],
                                                                                 item[i])
                        # Square Color ---------------------------------------
                        if col_as_name[col_names.index("squarecolor")] and item[i][headers.index("squarecolor")] != '':
                            point.style.iconstyle.scale = 0
                        # Icon color ----------------------------------------
                        if col_as_name[col_names.index("iconcolor")] and item[i][headers.index("iconcolor")] != '':
                            point.style.iconstyle.color = self.__color_translate(item[i][headers.index("iconcolor")])
                        # Icon scale / size ---------------------------------
                        if col_as_name[col_names.index("iconscale")] and item[i][headers.index("iconscale")] != '':
                            point.style.iconstyle.scale = item[i][headers.index("iconscale")]
                        # Ballon Description ----------------------------------
                        if col_as_name[col_names.index("description")]:
                            try:  # hours
                                point.description = self.__formated_time(item[i][headers.index("description")])
                            except ValueError:
                                point.description = str(item[i][headers.index("description")])
                        # Icon heading / inclination ------------------------
                        if col_as_name[col_names.index("iconheading")] and item[i][headers.index("iconheading")] != '':
                            heading = item[i][headers.index("iconheading")]
                            if type(heading) == float or type(heading) == int:
                                point.style.iconstyle.heading = heading
                            else:
                                heading = item[i][headers.index("icon")]
                                if not next_coords:
                                    next_coords = coords
                                point.style.iconstyle.heading = self.__icon_heading(coords, next_coords, heading)
                        # Color and data to build the line ------------------
                        if col_as_name[col_names.index("linestringcolor")] and \
                                item[i][headers.index('linestringcolor')] != '':
                            line = [(coords[0][0], coords[0][1])]
                            line_color = self.__color_translate(item[i][headers.index("linestringcolor")])

                            if i < len(item) - 1:  # Next coordinate (line)
                                if not next_coords:
                                    next_coords = coords
                                line.append((next_coords[0][0], next_coords[0][1]))
                                lin = folder.newlinestring(coords=line)
                                lin.style.linestyle.color = line_color
                                lin.style.linestyle.width = 2  # 2 pixels
                    # Polygon -------------------------------------------
                    if col_as_name[col_names.index("polygon")] and item[i][headers.index("polygon")] != '':
                        try:  # hours
                            description = self.__formated_time(item[i][headers.index("description")])
                        except TypeError:
                            description = str(item[i][headers.index("description")])
                        try:
                            name = self.__formated_time(item[i][2])
                        except TypeError:
                            try:
                                name = str(item[i][2])
                            except TypeError:
                                name = item[i][2]
                        pol = folder.newpolygon(name=name)
                        pol.altitudemode = simplekml.AltitudeMode.relativetoground
                        radius = float(item[i][headers.index("polygon")])
                        color = item[i][headers.index("polygoncolor")]
                        pol_color = ColornameToKml(color).get_color()
                        altitude = float(item[i][headers.index("polygonaltitude")])
                        azimute = float(item[i][headers.index("polygonazimute")])
                        if col_as_name[col_names.index("polygonamplitude")]:
                            amplitude = float(item[i][headers.index("polygonamplitude")])
                        else:
                            amplitude = 60.0
                        pol_points = self.__polygon(float(coords[0][0]),
                                                    float(coords[0][1]), azimute,
                                                    radius, altitude, amplitude)
                        pol.outerboundaryis = pol_points
                        pol.style.balloonstyle.text = self.__point_description(headers,
                                                                               [item[i][headers.index(add)]], item[i])
                        pol.style.linestyle.color = pol_color
                        pol.style.polystyle.color = simplekml.Color.changealphaint(100, pol_color)
                    # Square -------------------------------------------
                    if col_as_name[col_names.index("squarecolor")] and \
                            item[i][headers.index("squarecolor")] != '':
                        description = str(item[i][headers.index("description")])
                        sqr = folder.newpolygon(name=description)
                        sqr.altitudemode = simplekml.AltitudeMode.relativetoground
                        color = item[i][headers.index("squarecolor")]
                        sqr_color = ColornameToKml(color).get_color()
                        altitude = float(item[i][headers.index("squarealtitude")])
                        sqr_lat_original = item[i][headers.index("squarelatitude")]
                        sqr_lon_original = item[i][headers.index("squarelongitude")]
                        sqr_coords = Coordinates(str(sqr_lat_original), str(sqr_lon_original)).convert()
                        sqr_lat = sqr_coords[0]
                        sqr_lon = sqr_coords[1]
                        sqr_points = [[coords[0][0], coords[0][1], altitude],
                                      [sqr_lon, coords[0][1], altitude],
                                      [sqr_lon, sqr_lat, altitude],
                                      [coords[0][0], sqr_lat, altitude],
                                      [coords[0][0], coords[0][1], altitude]]
                        sqr.outerboundaryis = sqr_points
                        sqr.style.balloonstyle.text = self.__point_description(headers,
                                                                               [item[i][headers.index(add)]],
                                                                               item[i])
                        sqr.style.linestyle.color = sqr_color
                        sqr.style.polystyle.color = simplekml.Color.changealphaint(100, sqr_color)

        return kml, self.images_list

    def save_kmz(self, kml):
        """
        kml object -> None

        Turns Kml() object in a KMZ file an saves it in the disk.
        """
        path = self.file_name[:self.file_name.rindex(os.sep)]
        path_1 = self.file_name[self.file_name.rindex(os.sep) + 1:self.file_name.rfind('.')]
        kmzs = [x for x in os.listdir(path) if x[-4:] == '.kmz' and x[:-12] == path_1]

        if len(kmzs) > 0:
            kmzs.sort()
            version = str(round(float(kmzs[-1][-7:-4]) + .1, 2))
            kml.savekmz(self.file_name[:self.file_name.rfind('.')] + "_ver-" + version + ".kmz")
        else:
            version = "_ver-0.1.kmz"
            kml.savekmz(self.file_name[:self.file_name.rfind('.')] + version)

    def __point_description(self, headers, append_data_columns_to_description, data):
        """
        list, list, list -> str

        Builds a HTML string to be used in the description/ descriptive
        balloon of a Kml.point()
        """
        new_data = data[:]  # to manipulate a data copy
        new_headers = headers[:]  # to manipulate a headers copy
        # the column "AppendDataColumnsToDescription" items
        items = [x.split(',') for x in append_data_columns_to_description][0]
        f_items = [x.lower().strip() for x in items]  # formated items

        # Remove latitude and longitude and built a new Coordenadas
        if "latitude" and "longitude" in f_items:
            coordenadas = str(new_data[new_headers.index("latitude")]) + ", " + \
                          str(new_data[new_headers.index("longitude")])
            data_i = min(new_headers.index("latitude"), new_headers.index("longitude"))
            new_data.pop(new_headers.index("latitude"))
            new_headers.pop(new_headers.index("latitude"))
            new_data.pop(new_headers.index("longitude"))
            new_headers.pop(new_headers.index("longitude"))
            new_data.insert(data_i, coordenadas)
            new_headers.insert(data_i, "coordinates")
            f_items_i = min(f_items.index("latitude"), f_items.index("longitude"))
            f_items.pop(f_items.index("latitude"))
            f_items.pop(f_items.index("longitude"))
            f_items.insert(f_items_i, "coordinates")

        # The data indexes of the elements in AppendDataToColToDescr
        indexes = [new_headers.index(x.lower().strip()) for x in f_items if x.lower().strip() in new_headers]
        # Add to the indexes the Description column --------------------
        indexes.insert(0, new_headers.index("description"))

        # Format Dates and Times ---------------------------------------
        if "name" in f_items and type(new_data[new_headers.index("name")]) is tuple:  # Excel time format
            i = new_headers.index("name")
            new_data.insert(i, self.__formated_time(new_data.pop(i)))
        if "data" in f_items:
            i = new_headers.index("data")
            new_data.insert(i, self.__formated_date(new_data.pop(i)))
        if "hora" in f_items:
            i = new_headers.index("hora")
            new_data.insert(i, self.__formated_time(new_data.pop(i)))
        if "duracao" in f_items:
            i = new_headers.index("duracao")
            new_data.insert(i, self.__formated_time(new_data.pop(i)))
        if "duração" in f_items:
            i = new_headers.index("duração")
            new_data.insert(i, self.__formated_time(new_data.pop(i)))
        if "cellfix time" in f_items:
            i = new_headers.index("cellfix time")
            new_data.insert(i, self.__formated_time(new_data.pop(i)))

        # Capitalize titles --------------------------------------------
        pt = [x.capitalize() for x in f_items]
        # Photos -------------------------------------------------------
        if "foto" in f_items:
            return self.__point_description_foto(new_data, indexes, pt)

        # HTML Format --------------------------------------------------
        tags = {0: '<BalloonStyle><text>',
                1: '<table><tr><td colspan="2" align="center">',
                2: '</td></tr>',
                3: '<tr style="background-color:lightgreen"><td align="left">',
                4: '<tr><td>',
                5: '</td></tr>',
                6: '<td>',
                7: '</td>',
                8: '</table>',
                9: '</text></BalloonStyle>'}

        # Return build -------------------------------------------------
        try:
            title = self.__formated_time(new_data[indexes[0]])
        except:
            title = str(new_data[indexes[0]])

        head = tags[0] + tags[1] + str(title) + tags[2]
        body = []
        tail = tags[8] + tags[9]

        for i in range(1, len(indexes)):
            if i % 2 != 0:
                body.append(tags[3] + pt[i - 1] + ": " + tags[7] +
                            tags[6] + str(new_data[indexes[i]]) + tags[5])
            else:
                body.append(tags[4] + pt[i - 1] + ": " + tags[7] +
                            tags[6] + str(new_data[indexes[i]]) + tags[5])

        # Return -------------------------------------------------------
        return head + ''.join(body) + tail

    def __point_description_foto(self, data, indexes, titles):
        '''
        Build the descriptions / descriptive balloons witch have photos.
        '''
        path = os.getcwd() + os.sep + "fotos" + os.sep  # photos path
        original_path = os.getcwd()
        msg = "O nome da foto no Excel difere do nome da foto na pasta 'fotos'"

        # HTML Format --------------------------------------------------
        tags = {0: '<![CDATA[<BalloonStyle><text>',
                1: '<table width="400" border="0" cellspacing="5" \
                    cellpadding="3"><tr><td colspan="2" align="center">\
                    <font color="0000ff"><b>',
                2: '</b></font></td></tr>',
                3: '<tr style="background-color:lightgreen">\
                    <td align="center">',
                4: '<tr><td colspan="2" align="center"></h3>',
                5: '</h3></td></tr>',
                6: '<tr><td colspan="2" align="center">',
                7: '</td></tr>',
                8: '</table>',
                9: '\n<img src="',
                10: '" alt="foto" width="400" height="280">\n</br>',
                11: '<tr><td><hr></td></tr>\
                    <tr style="backgound-color:lightgreen" ><td></td></tr>',
                12: '</text></BalloonStyle>]]>'}

        # Build --------------------------------------------------------
        head = tags[0] + tags[1] + str(data[indexes[0]]) + tags[2] + tags[6] + tags[7]
        body = []
        tail = tags[8] + tags[12]

        for i in range(len(titles)):
            if "foto" in titles[i].lower():
                os.chdir(path)
                if not os.path.isfile(data[indexes[i + 1]]):
                    os.chdir(original_path)
                    logfile = open('error.log', 'a')
                    logfile.write(msg)
                    logfile.close()
                    os.startfile('error.log')
                    try:
                        os.system('taskkill /F /T /IM xls2kmz.exe')
                    except:
                        os.system('taskkill /F /T /IM pythonw.exe')
                # --------------------------------------------------------------
                else:
                    os.chdir(original_path)
                    body.append(tags[6] + tags[9] + 'files/' + (data[indexes[i + 1]]) + tags[10] + tags[7])
                    if path + (data[indexes[i + 1]]) not in self.images_list:
                        self.images_list.append(path + (data[indexes[i + 1]]))
            elif "descrição" and "descricao" in titles[i].lower():
                body.append(tags[6] + (data[indexes[i + 1]]) + tags[7] + tags[11])
            else:
                body.append(tags[3] + (data[indexes[i + 1]]) + tags[7])

        # Return -------------------------------------------------------
        return head + ''.join(body) + tail

    def __formated_date(self, xldate):
        """
        tuple -> str

        xldade is a tuple with the EXEL datetime value.

        Formates the EXEL cell values of datetime type.
        Puts a "-" separating days from months and months from years.
        Has as parameter is passed a tuple, xldate, which has the
        datetime value of the EXEL cell.
        """
        pattern = '{0:02d}'  # two numeric places

        if type(xldate) is not tuple:
            return str(xldate)

        return pattern.format(xldate[2]) + '-' + pattern.format(xldate[1]) + '-' + pattern.format(xldate[0])

    def __formated_time(self, xldate):
        """
        tuple -> str

        xldade is a tuple with the EXEL datetime value.

        Formates the EXEL cell values of datetime type.
        Puts a  ":" between hours and minutes, and minutes and seconds.
        """
        pattern = '{0:02d}'  # two numeric places

        if xldate == (0, 0, 0, 0, 0, 0) or xldate == 0:  # 0 duration
            return "00:00:00"
        if type(xldate) is not tuple:
            return str(xldate)

        return pattern.format(xldate[3]) + ':' + pattern.format(xldate[4]) + ':' + pattern.format(xldate[5])

    def __icon_heading(self, coords, next_coords, icon):
        """
        list, list, int -> float

        Calculates and returns in grades the icon direction of the
        Kml().points.
        Coords are the actual coordinates.
        next_coords are the next coordinates.
        icon is the icon number.

        In a right angled triangle, the hypotenuse is is the side
        opposite to the 90 degrees angle, the opposite is the side
        opposite to the angle we want to find, and, the adjacent, is
        the side who joins the angle we want to find to the 90
        degree angle.
        To find that angle we need to use the Inverse Tangent or
        ArcTangent

        The Tangent of the angle ø is:
            tan(ø) = Opposite / Adjacent
        So, the inverse Tangent is:
            tan^-1(Opposite / Adjacent) = ø

        See: http://www.mathsisfun.com/algebra/trig-inverse-sin-cos-tan.html
        """
        adjac = float(next_coords[0][1]) - float(coords[0][1])
        oppos = float(next_coords[0][0]) - float(coords[0][0])
        angle = 0.0  # If the first adjac is 0

        if adjac != 0.0:  # avoid ZeroDivisionError
            angle = math.atan(oppos / adjac)
        else:
            if oppos < 0:
                return 90.0
            else:
                return - 90.0
        routable_icons = [38, 106, 338, 350, 1000]
        if icon in routable_icons and adjac < 0:  # difference of negative longitude
            return math.degrees(angle)

        if icon in routable_icons and adjac >= 0:  # difference of positive longitude
            return math.degrees(angle) - 180

        return 0.0  # other icons

    def __color_translate(self, color):
        """
        str -> str

        In KML, the values for the color and opacity (alpha) are
        expressed in hexadecimal notation. The range of values for any
        color are 0 to 255 (00 to FF). For the alpha 00 is totally
        transparent and FF is totally opaque.
        The order of the expression are AABBGGRR, where AA=alpha,
        BB=blue, GG=green, and RR=red.
        """
        return ColornameToHex(color).get_abgr()

    def __polygon(self, latitude, longitude, azimute, radius, altitude,
                  amplitude):
        """
        float, float, int, float, float, float -> str

        Construts a polygon (quasi triangle) given the initial point
        (latitude and longitude), the direction (azimute), the radius,
        the altitude and the amplitude (open degree).
        """
        circle_points = self.__spoints(latitude, longitude, radius, altitude, 72, 0)

        if int(amplitude) == 360:
            return circle_points
        else:
            azi_point = int(round(azimute / 10 * 2))
            n_points = int(round(amplitude / 10 * 2))
            origin = (latitude, longitude, altitude)
            triangle_points = [origin]
            reverse_choose_points = []

            for i in range(int(round(n_points / 2))):
                reverse_choose_points.append(circle_points[(azi_point - i) % 73])

            for pt in reversed(reverse_choose_points):
                triangle_points.append(pt)

            for i in range(int(round(n_points / 2))):
                triangle_points.append(circle_points[(azi_point + i) % 73])

            triangle_points.append(origin)

            return triangle_points

    def __toEarth(self, p, altitude):
        if p[0] == 0.0:
            longitude = math.pi / 2.0
        else:
            longitude = math.atan(p[1] / p[0])
        colatitude = math.acos(p[2])
        latitude = (math.pi / 2.0 - colatitude)

        # select correct branch of arctan
        if p[0] < 0.0:
            if p[1] <= 0.0:
                longitude = -(math.pi - longitude)
            else:
                longitude = math.pi + longitude

        DEG = 180.0 / math.pi

        return [longitude * DEG, latitude * DEG, altitude]

    def __toCart(self, longitude, latitude):
        """
        convert long, lat IN RADIANS to (x,y,z)

        spherical coordinate use "co-latitude", not "latitude"
        latiude = [-90, 90] with 0 at equator
        co-latitude = [0, 180] with 0 at north pole
        """
        theta = longitude
        phi = math.pi / 2.0 - latitude

        return [math.cos(theta) * math.sin(phi), math.sin(theta) * math.sin(phi), math.cos(phi)]

    def __spoints(self, lon, lat, meters, altitude, n, offset=0):
        """
        __spoints -- get raw list of points in long,lat format

        meters: radius of polygon
        n: number of sides
        offset: rotate polygon by number of degrees

        Returns a list of points comprising the object
        """
        RAD = math.pi / 180.0  # constant to convert to radians
        MR = 6378.1 * 1000.0  # Mean Radius of Earth, meters
        offsetRadians = offset * RAD
        # compute longitude degrees (in radians) at given latitude
        r = (meters / (MR * math.cos(lat * RAD)))

        vec = self.__toCart(lon * RAD, lat * RAD)
        pt = self.__toCart(lon * RAD + r, lat * RAD)
        pts = []

        for i in range(0, n):
            pts.append(self.__toEarth(self.__rotPoint(vec, pt, offsetRadians + (2.0 * math.pi / n) * i), altitude))

        pts.append(pts[0])  # connect to starting point exactly

        return pts

    def __rotPoint(self, vec, pt, phi):
        '''
        rotate point pt, around unit vector vec by phi radians
        http://blog.modp.com/2007/09/rotating-point-around-vector.html
        '''
        # remap vector for sanity
        (u, v, w, x, y, z) = (vec[0], vec[1], vec[2], pt[0], pt[1], pt[2])

        a = u * x + v * y + w * z
        d = math.cos(phi)
        e = math.sin(phi)

        return [(a * u + (x - a * u) * d + (v * z - w * y) * e),
                (a * v + (y - a * v) * d + (w * x - u * z) * e),
                (a * w + (z - a * w) * d + (u * y - v * x) * e)]


class MdkControl(object):
    """
    Turns a list into a Mdk() object and saves in the disk as a MKD file.
    """

    def __init__(self, data_list, filename):
        """
        """

        self.data_list = data_list
        self.filename = filename
        self.markupdata = et.Element('MarkupData')
        self.__defaultSubElements()
        self.routes = et.SubElement(self.markupdata, 'Routes')
        self.tree = None

    def build_mkd(self):
        """
        """
        for item in self.data_list:
            route = et.SubElement(self.routes, 'Route')
            et.SubElement(route, 'Name').text = str(item[0][0])  # sheetname
            et.SubElement(route, 'GUID').text = self.__getGUID()
            et.SubElement(route, 'Color A="255" R="0" G="0" B="0"')
            pos = et.SubElement(route, 'Positions')
            if len(item) > 1:
                headers = [x.lower() for x in item[1]]  # column titles
                # Optional column names to build a point-----------------
                col_names = ["icon", "size", "description", "rotation", "appenddatacolumnstodescription", "size"]
                # See what optional names are in the EXEL data ----------
                col_as_name = [True if x in headers else False for x in col_names]
                #  POINT BUILD ##########################################
                for i in range(2, len(item)):
                    rt_pos = et.SubElement(pos, 'RoutePosition')
                    # Coordinates ---------------------------------------
                    lat1 = str(item[i][0])
                    lon1 = str(item[i][1])
                    coords = Coordinates(lon1, lat1).convert()
                    #  Point --------------------------------------------
                    if col_as_name[col_names.index('icon')] and item[i][headers.index('icon')] != '':
                        t = str(item[i][headers.index('icon')])
                        et.SubElement(rt_pos, 'Icon').text = t

                    et.SubElement(rt_pos, 'Name').text = str(item[i][2])

                    if col_as_name[col_names.index('description')] and item[i][headers.index('description')] != '':
                        add = 'appenddatacolumnstodescription'
                        text = self.__point_description(headers, [item[i][headers.index(add)]], item[i])
                        et.SubElement(rt_pos, 'Description').text = text

                    et.SubElement(rt_pos, 'ShowLabel').text = 'true'

                    if col_as_name[col_names.index('rotation')] and item[i][headers.index('rotation')] != '':
                        t = item[i][headers.index('rotation')]
                        et.SubElement(rt_pos, 'Rotation').text = t
                    else:
                        et.SubElement(rt_pos, 'Rotation').text = '0'

                    if col_as_name[col_names.index('size')] and item[i][headers.index('size')] != '':
                        t = str(item[i][headers.index('size')])
                        et.SubElement(rt_pos, 'Size').text = t
                    else:
                        et.SubElement(rt_pos, 'Size').text = '24'

                    geopoint = et.SubElement(rt_pos, 'GeoPoint')
                    et.SubElement(geopoint, 'Lat').text = str(coords[1])
                    et.SubElement(geopoint, 'Lon').text = str(coords[0])

        return et.ElementTree(self.markupdata)

    def save_mkd(self, tree):
        """
        """

        path = self.filename[:self.filename.rindex('\\')]
        path1 = self.filename[self.filename.rindex("\\") + 1:self.filename.rfind('.')]

        mkds = [x for x in os.listdir(path) if x[-4:] == '.mkd' and x[:-12] == path1]

        if len(mkds) > 0:
            mkds.sort()
            version = str(round(float(mkds[-1][-7:-4]) + .1, 2))
            tree.write(self.filename[:self.filename.rfind('.')] + "_ver-" + version + ".mkd")
        else:
            version = "_ver-0.1.mkd"
            tree.write(self.filename[:self.filename.rfind('.')] + version)

    def __defaultSubElements(self):
        """
        """
        et.SubElement(self.markupdata, 'FileVersion').text = '1.1'

        for e in ['Rectangles', 'Circles', 'Arcs', 'Polygons', 'Lines', 'Placemarks']:
            et.SubElement(self.markupdata, e)

    def __setGUID(self):
        """
        Returns a random hexadecimal value
        """

        start = 1000000000000000000
        stop = 9999999999999999999

        return str(hex(random.randrange(start, stop)))[2:]

    def __getGUID(self):
        """
        Returns a random hexadecimal value
        """

        return self.__setGUID()

    def __point_description(self, headers, appenddata, data):
        """
        """

        new_data = data[:]  # to manipukate a data copy
        new_headers = headers[:]  # to manipulate an headers copy
        # the column "AppendDataColumnsTo Description" items
        items = [x.split(',') for x in appenddata][0]
        f_items = [x.lower().strip() for x in items]  # formated items

        # the data indexes of the elements in AppendDataToColToDescr
        indexes = [new_headers.index(x.lower().strip()) for x in f_items if x.lower().strip() in new_headers]
        # add to the indexes the Description column
        indexes.insert(0, new_headers.index('description'))

        pt = [x.capitalize() for x in f_items]  # Capitalize titles

        text = str(new_data[indexes[0]]) + '\n'
        for i in range(1, len(indexes)):
            text += pt[i - 1] + ': ' + str(new_data[indexes[i]]) + '\n'

        return text


class MotherControl(object):
    """
    Calls the controls xls.py, kml.py and create_kmz.py, and makes
    instances of XlsControl(), KmlControl() and CreateKMZ() objects.
    """

    def __init__(self, ficheiro, original_working_dir):
        """
        str -> object MotherControl() object

        file_name is a string, the name of the EXEL file, which we want
        to make a KMZ file from.
        """
        self.ficheiro = ficheiro
        self.original_working_dir = original_working_dir

    def excel_to_mkd(self):
        """
        None -> None

        """
        excel = ExcelSplitter(self.ficheiro)
        excel_list = excel.splitter()
        # -------------------------------
        if type(excel_list) != list:
            return excel_list

        for excel in excel_list:
            xls = XlsControl(excel)
            data_list = xls.read_exel()
            # --------------------------
            if type(data_list) != list:
                return data_list
            # --------------------------
            mkd_list = MdkControl(data_list, os.path.abspath(excel))
            mkd = mkd_list.build_mkd()
            mkd_list.save_mkd(mkd)

        # Voltar à working directory original
        os.chdir(self.original_working_dir)

    def excel_to_kml(self):
        """
        None -> None

        Makes an instance of a XlsControl() object, by one attribute of
        that object build a list with all the data from the EXEL cells.
        Then, makes an instance of a KmlControl() object, by one of his
        attributes, makes an Kml() object, which is passed as an
        argument of another KmlControl() attribute to save a KMZ file in
        the drive.
        """
        global kml
        excel = ExcelSplitter(self.ficheiro)
        excel_list = excel.splitter()
        # -------------------------------
        if type(excel_list) != list:
            return excel_list
        # -------------------------------
        kmzs_list = []

        for excel in excel_list:
            xls = XlsControl(excel)
            data_list = xls.read_exel()
            # --------------------------
            if type(data_list) != list:
                return data_list
            # ---------------------------
            kml_list = KmlControl(data_list, os.path.abspath(excel))
            kml = kml_list.build_kml()
            kml_list.save_kmz(kml[0])

            path = os.path.abspath(os.path.dirname(excel))
            path_1 = os.path.abspath(os.path.splitext(excel)[0])
            kmzs = [x for x in os.listdir(path) if x[-4:] == '.kmz']
            kmzs.sort()
            kmz_file = path + os.sep + kmzs[-1]
            kmzs_list.append(kmz_file)
            kmz = CreateKMZ(kmz_file, kml[1])
            kmz.rebuild_kmz()

        # Aqui tenho de ver se já lá há kmz e por '_ver-0.n'
        original_dir_path = os.path.abspath(os.path.dirname(self.ficheiro))
        abs_original_file_path = os.path.abspath(
            os.path.splitext(self.ficheiro)[0])
        original_file_path = abs_original_file_path[
                             abs_original_file_path.rfind('\\') + 1:]
        kmzs = [x for x in os.listdir(original_dir_path) if x[-4:] == '.kmz' and x[:-12] == original_file_path]
        kmzs.sort()
        file_name = ''
        if len(kmzs) == 0:
            file_name = self.ficheiro[:self.ficheiro.rfind('.')] + \
                        '_ver-0.1.kmz'
        else:
            version = str(round(float(kmzs[-1][-7:-4]) + .1, 2))
            file_name = self.ficheiro[:self.ficheiro.rfind('.')] + '_ver-' + version + '.kmz'

        if os.path.exists(
                os.path.abspath(os.path.dirname(self.ficheiro)) + '\\Temp'):
            final = KmlJoiner(kmzs_list, file_name)
            final.build_new_kmz()
        else:
            try:
                path = self.ficheiro[:self.ficheiro.rindex('/')]
                path_1 = self.ficheiro[self.ficheiro.rindex('/') + 1:self.ficheiro.rfind('.')]
                kmzs = [x for x in os.listdir(path) if x[-4:] == '.kmz' and x[:-12] == path_1]
                kmzs.sort()
                kmz_file = path + os.sep + kmzs[-1]
                kmz = CreateKMZ(kmz_file, kml[1])
                kmz.rebuild_kmz()
            except:
                pass

        # LIMPAR A TEMP
        if os.path.exists(
                os.path.abspath(os.path.dirname(self.ficheiro)) + '\\Temp'):
            os.chdir(
                os.path.abspath(os.path.dirname(self.ficheiro)) + '\\Temp')
            clean = TempCleaner(os.getcwd())
            clean.clean()

        # Voltar à working directory original
        os.chdir(self.original_working_dir)


class CreateKMZ(object):
    """
    Control to create a KMZ file from a KML one
    """

    def __init__(self, kmz_file, images_list):
        self.kmz_file = ''
        for char in kmz_file:
            if char == '/':
                self.kmz_file = self.kmz_file + os.sep
            else:
                self.kmz_file = self.kmz_file + char
        self.images = images_list

    def rebuild_kmz(self):
        zf = zipfile.ZipFile(self.kmz_file, "a")

        for image in self.images:
            # Relative Path to the Image
            try:
                zf.write(image, arcname='files/' + image[image.rfind('\\') + 1:])
            except:
                pass
        zf.close()


class Xls2kml(object):
    """
    Interface builted in Tkinter()
    """

    def __init__(self):
        """
        None -> None

        Builds the Tkinter window and all his elements.
        """
        # variables ----------------------------------------------------
        # log file
        open("error.log", "w").close()  # to open and clean the logfile
        logging.basicConfig(level=logging.DEBUG, filename='error.log')
        sys.stderr = LogFile('stderr')  # Redirect stderr
        self.original_working_dir = os.getcwd()  # original working dir
        self.master = Tk()  # Tk() object
        self.master.title('EXCEL to KMZ Transformer - 2.0.6')  # window name
        icons = os.getcwd() + os.sep + "icons" + os.sep  # path to icons
        foto_folder = os.getcwd() + os.sep + "fotos"  # path to fotos
        icon = icons + "compass.ico"
        if os.name == 'nt':
            self.master.iconbitmap(icon)  # window icon
        self.master.resizable(width=FALSE, height=FALSE)
        self.master.geometry("566x314")
        self.file_name = ""  # the name of the EXEL file
        self.last_dir = "C:/"
        # image to decorate the window
        self.img = ImageTk.PhotoImage(Image.open(icons + "excel-kmz.jpg"))
        # to use in frame, message, labels and buttons -----------------
        self.message = StringVar()
        self.message.set("\nSelect an EXCEL file")
        bg = "gray25"
        bg1 = "dark orange"
        fc = "white smoke"
        font = ("Helvetica", "8", "bold")
        font1 = ("Helvetica", "10", "bold")
        text0 = " -- "
        text1 = " Boris & Vladimir Software "
        text = text0 + text1 + text0

        # Menu ---------------------------------------------------------
        self.menu = Menu(self.master)
        self.master.config(menu=self.menu)
        filemenu = Menu(self.menu)
        self.menu.add_cascade(label="File", menu=filemenu)
        filemenu.add_command(label="Quit", command=self.__callback_3)
        filemenu.add_command(label='Images Folder', command=lambda: (self.__open_folder(foto_folder)))

        self.openGE = BooleanVar()
        self.openGE.set(False)
        optionsmenu = Menu(self.menu)
        self.menu.add_cascade(label="Options", menu=optionsmenu)
        optionsmenu.add_checkbutton(label="Do not start Google Earth",
                                    onvalue=True, offvalue=False, variable=self.openGE)
        docsmenu = Menu(self.menu)
        docs = ["docs\manual.pdf", "docs\icons.pdf", "docs\colors.pdf",
                "docs\GPS.xlsx", "docs\GPS.kmz", "docs\Celulas.xlsx",
                "docs\Celulas.kmz", "docs\Foto.xlsx", "docs\Foto.kmz",
                "docs\Quadrado.xls", "docs\Quadrado.kmz"]
        self.menu.add_cascade(label="Documentation", menu=docsmenu)
        docsmenu.add_command(label="Manual", command=lambda: (self.__open_file(docs[0])))
        docsmenu.add_command(label="Icons", command=lambda: (self.__open_file(docs[1])))
        docsmenu.add_command(label="Colors", command=lambda: (self.__open_file(docs[2])))

        exemplemenu = Menu(docsmenu)
        docsmenu.add_cascade(label="Examples", menu=exemplemenu)

        gpsmenu = Menu(exemplemenu)
        exemplemenu.add_cascade(label="GPS", menu=gpsmenu)
        gpsmenu.add_command(label="Excel", command=lambda: (self.__open_file(docs[3])))
        gpsmenu.add_command(label="Google Earth", command=lambda: (self.__open_file(docs[4])))

        cellmenu = Menu(exemplemenu)
        exemplemenu.add_cascade(label="Mobile Cells", menu=cellmenu)
        cellmenu.add_command(label="Excel", command=lambda: (self.__open_file(docs[5])))
        cellmenu.add_command(label="Google Earth", command=lambda: (self.__open_file(docs[6])))

        fotomenu = Menu(exemplemenu)
        exemplemenu.add_cascade(label="Fotos", menu=fotomenu)
        fotomenu.add_command(label="Excel", command=lambda: (self.__open_file(docs[7])))
        fotomenu.add_command(label="Google Earth", command=lambda: (self.__open_file(docs[8])))

        squaremenu = Menu(exemplemenu)
        exemplemenu.add_cascade(label="Square", menu=squaremenu)
        squaremenu.add_command(label="Excel", command=lambda: (self.__open_file(docs[9])))
        squaremenu.add_command(label="Google Earth", command=lambda: (self.__open_file(docs[10])))

        helpmenu = Menu(self.menu)
        self.menu.add_cascade(label='Help', menu=helpmenu)
        helpmenu.add_command(label="About", command=self.__about)
        helpmenu.add_command(label="View Log", command=lambda: (self.__open_file("error.log")))

        # Frame to suport butons, labels and separators ----------------
        self.f = Frame(self.master, bg=bg)
        self.f.pack_propagate(0)  # don't shrink
        self.f.pack(side=BOTTOM, padx=0, pady=0)

        # Message and Labels -------------------------------------------
        self.l1 = Message(
            self.f, bg=bg1, bd=5, fg=bg, textvariable=self.message,
            font=("Helvetica", "13", "bold italic"), width=500).grid(
            row=0, columnspan=6, sticky=EW, padx=5, pady=5)
        self.l2 = Label(
            self.f, image=self.img, fg=bg).grid(
            row=1, columnspan=6, padx=0, pady=2)
        self.l6 = Label(
            self.f, text=text, font=font1, bg=bg, fg=bg1).grid(
            row=3, column=3, columnspan=2, sticky=EW, pady=5)

        # Buttons ------------------------------------------------------
        self.b0 = Button(
            self.f, text="Open EXCEL...", command=self.__callback, width=10,
            bg="forest green", fg=fc, font=font).grid(row=3, column=0, padx=5, sticky=W)
        self.b1 = Button(
            self.f, text="Save KMZ", command=self.__callback_2, width=10,
            bg="DodgerBlue3", fg=fc, font=font).grid(row=3, column=1, padx=5, sticky=W)
        self.b2 = Button(
            self.f, text="Quit", command=self.__callback_3, width=10,
            bg="orange red", fg=fc, font=font).grid(row=3, column=5, sticky=E, padx=5)
        self.b3 = Button(
            self.f, text="Save MKD", command=self.__callback_4, width=10,
            bg="DodgerBlue4", fg=fc, font=font).grid(row=3, column=2, padx=5, sticky=W)

        # Progressbar --------------------------------------------------
        self.s = ttk.Style()
        # themes: winnative, clam, alt, default, classic, vista, xpnative
        if os.name == 'nt':
            self.s.theme_use('winnative')
        self.s.configure("red.Horizontal.TProgressbar", foreground='green', background='forest green')
        self.pb = ttk.Progressbar(self.f, orient='horizontal', mode='determinate',
                                  style="red.Horizontal.TProgressbar")
        self.pb.grid(row=2, column=0, columnspan=6, padx=5, pady=5, sticky=EW)

        # Mainloop -----------------------------------------------------
        self.master.mainloop()

    def __callback(self):  # "Abrir EXEL..." button handler ------------
        """
        None -> None

        Opens a new window (filedialog.askopenfilename) to choose the
        EXCEL file that is necessary to make the KMZ file.
        """
        title = 'Select an Excel file'
        msg = 'EXCEL loaded in memory\nTransform it to KMZ/MKD'
        self.file_name = filedialog.askopenfilename(title=title, initialdir=self.last_dir)
        self.last_dir = self.file_name[:self.file_name.rfind('/')]

        if self.file_name[self.file_name.rfind('.') + 1:] != 'xls' and \
                self.file_name[self.file_name.rfind('.') + 1:] != 'xlsx':
            msg = self.file_name + ' Is not a valid Excel file!'
        self.message.set(msg)

    def __callback_2(self):  # "Gravar KMZ" button handler ---------------
        """
        None -> None

        Calls the function self.__threat("kmz")
        """
        sleep(1)
        msg = 'EXCEL loaded in memory\nTransform it to KMZ/MKD'
        if self.message.get() != msg:
            self.message.set("\nFirst choose an Excel file")
        else:
            self.message.set("\nProcessing...")
            self.master.update_idletasks()
            sleep(1)
            self.__threads("kmz")

    def __callback_3(self):  # "Sair" button handler ---------------------
        """
        None -> None

        Kills the window
        """
        self.master.destroy()

    def __callback_4(self):  # "Gravar MKD" button handler ---------------
        """
        None -> None

        Calls the function self.__threads("mkd")
        """
        sleep(1)
        msg = 'EXCEL loaded in memory\nTransform it to KMZ/MKD'
        if self.message.get() != msg:
            self.message.set("\nFirst choose an Excel file")
        else:
            self.message.set("\nProcessing...")
            self.master.update_idletasks()
            sleep(1)
            self.__threads("mkd")

    def __threads(self, mkd_or_kmz):
        """
        str -> MyTread() objects

        mkd_or_kmz - a string to choose between kmz or mdk

        Creates two threads to run at the same time the functions:
        self.__create_kmz() or self.__crerate_mkd()
        self.__progressbar()
        """
        if mkd_or_kmz == "mkd":
            funcs = [self.__create_mkd, self.__progressbar]
        else:
            funcs = [self.__create_kmz, self.__progressbar]
        threads = []
        nthreads = list(range(len(funcs)))

        for i in nthreads:
            t = MyThread(funcs[i], (), funcs[i].__name__)
            threads.append(t)

        for i in nthreads:
            threads[i].start()

    def __create_mkd(self):
        """
        None -> None

        Calls the excel_to_mkd() attribute from the MotherControl() class
        """
        mkd = MotherControl(
            self.file_name, self.original_working_dir).excel_to_mkd()
        if type(mkd) == str:
            self.message.set(mkd)
            self.pb.stop()
            self.master.update_idletasks()
        else:
            sleep(2)
            self.pb.stop()
            self.master.update_idletasks()
        self.message.set("\nMKD saved with success")
        sleep(2)
        self.master.update_idletasks()

    def __create_kmz(self):
        """
        None -> None

        Calls the excel_to_kml() atribute from MotherControl() class
        And when it returns, calls self.__open_Google_Earth()
        """
        kmz = MotherControl(
            self.file_name, self.original_working_dir).excel_to_kml()
        if type(kmz) == str:
            self.message.set(kmz)
            self.pb.stop()
            self.master.update_idletasks()
        else:
            sleep(2)
            self.pb.stop()
            self.master.update_idletasks()
            self.__open_Google_Earth()

    def __open_Google_Earth(self):
        """
        None -> None

        Opens the made KMZ file in Google Earth
        """
        sleep(1)
        self.master.update_idletasks()
        msg = "KMZ saved with success.\nOpening Google Earth..."
        if not self.openGE.get():
            self.message.set(msg)
        else:
            self.message.set("\nKMZ saved with success.\n")
        sleep(2)
        self.master.update_idletasks()
        path = self.file_name[:self.file_name.rindex('/')]
        path_1 = self.file_name[self.file_name.rindex('/') +
                                1:self.file_name.rfind('.')]
        kmzs = [x for x in os.listdir(path) if x[-4:] == '.kmz' and x[:-12] == path_1]
        kmzs.sort()
        try:
            if not self.openGE.get():
                os.startfile(path + os.sep + kmzs[-1])
                sleep(2)
            self.message.set("\nSelect an EXCEL file")
        except:
            msg = "Install Google Earth\nhttp://www.google.com/earth/"
            self.message.set(msg)
            self.master.update_idletasks()

    def __progressbar(self, ratio=0):
        """
        None -> None

        Starts the progressbar in the window
        """
        self.pb.start(50)

    def __about(self):
        """
        None -> None

        Associated with the Help Menu.
        Creates a new window with the "About" information
        """
        appversion = "2.0.6"
        appname = "EXCEL to KML Transformer"
        copyright = 14 * ' ' + '(c) 2013' + 12 * ' ' + \
            'SDATO - DP - UAF - GNR\n' + 34 * ' ' + "No Rights Reserved"
        licence = 18 * ' ' + 'http://opensource.org/licenses/GPL-3.0\n'
        contactname = "Nuno Venâncio"
        contactphone = "(00351) 969 564 906"
        contactemail = "venancio.gnr@gmail.com"

        message = "Version: " + appversion + 5 * "\n"
        message0 = "Copyleft: " + copyright + "\n" + "Licence: " + licence
        message1 = contactname + '\n' + contactphone + '\n' + contactemail

        icons = os.getcwd() + os.sep + "icons" + os.sep  # path to icons
        icon = icons + "compass.ico"

        tl = Toplevel(self.master)
        tl.configure(borderwidth=5)
        tl.title("About...")
        tl.iconbitmap(icon)
        tl.resizable(width=FALSE, height=FALSE)
        f1 = Frame(tl, borderwidth=2, relief=SUNKEN, bg="gray25")
        f1.pack(side=TOP, expand=TRUE, fill=BOTH)

        l0 = Label(f1, text=appname, fg="white", bg="gray25", font=('courier', 16, 'bold'))
        l0.grid(row=0, column=0, sticky=W, padx=10, pady=5)
        l1 = Label(f1, text=message, justify=CENTER, fg="white", bg="gray25")
        l1.grid(row=2, column=0, sticky=E, columnspan=3, padx=10, pady=0)
        l2 = Label(f1, text=message0, justify=LEFT, fg="white", bg="gray25")
        l2.grid(row=6, column=0, columnspan=2, sticky=W, padx=10, pady=0)
        l3 = Label(f1, text=message1, justify=CENTER, fg="white", bg="gray25")
        l3.grid(row=7, column=0, columnspan=2, padx=10, pady=0)

        button = Button(tl, text="Ok", command=tl.destroy, width=10)
        button.pack(pady=5)

    def __open_file(self, doc):
        try:
            os.startfile(doc)
        except:
            pass

    def __open_folder(self, folder):
        os.system('start explorer "' + folder + '"')


if __name__ == '__main__':
    Xls2kml()
